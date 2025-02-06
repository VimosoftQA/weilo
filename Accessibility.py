# 접근성 아이디 엑셀 파일을 불러와서 JSON 의 형태로 변환해준다.

import openpyxl as opx
from tqdm import tqdm
import json

excel = opx.load_workbook("WEILO_Accessibility_ID.xlsx")
all_sheets = excel.sheetnames

# 모든 시트의 Component Description 과 ID 가져오기
all_data = list()
for sheet_name in tqdm(all_sheets[1:]):
    sheet = excel[sheet_name]
    for row in sheet.iter_rows(min_row= 5, min_col=3, max_col=4):  # 필요한 행, 열 지정
        row_data = [cell.value for cell in row if cell.value is not None]
        all_data.append(row_data)

data_dictionary = dict()

for data in tqdm(all_data):
    if len(data) == 2 :
        data_dictionary[data[0]] = data[1]

print(data_dictionary) #179개

with open('accessibilityID.json', 'w') as f:
    json.dump(data_dictionary, f, ensure_ascii=False)



